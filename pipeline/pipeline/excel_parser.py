"""Parse the LAPIG Excel metadata file and expand dates/depths into item definitions."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Provider:
    name: str
    description: str
    roles: list[str]
    url: str


@dataclass
class ItemDefinition:
    """One STAC item to be generated."""

    item_id: str
    start_datetime: datetime
    end_datetime: datetime
    assets: dict[str, str]  # asset_key -> href


@dataclass
class CollectionRecord:
    """Parsed collection row with expanded item definitions."""

    id: str
    row_index: int
    dt_position: float | None
    catalog: str
    collection_class: str
    title: str
    description: str
    keywords: list[str]
    stac_url: str | None
    version: str | None
    doi: str | None
    citation: str | None
    license: str
    layer_unit: str | None
    scale: float | None
    offset: float | None
    spatial_aggregation: str | None
    contact_name: str | None
    contact_email: str | None
    provider_ids: list[str]
    gsd: float | None
    start_date: datetime | None
    end_date: datetime | None
    date_step: str | None  # raw value
    date_offset: str | None  # raw value
    date_unit: str | None
    date_style: str | None
    depth_list: list[str]
    main_url: str | None
    extra_urls: dict[str, str]  # url_1..url_13
    sld_urls: list[str]
    qml_urls: list[str]
    custom_bbox: list[float] | None = None  # [west, south, east, north]
    items: list[ItemDefinition] = field(default_factory=list)


@dataclass
class CatalogInfo:
    id: str
    title: str
    description: str


@dataclass
class ParseResult:
    catalog: CatalogInfo
    providers: dict[str, Provider]
    collections: list[CollectionRecord]


# ---------------------------------------------------------------------------
# ID extraction
# ---------------------------------------------------------------------------

_FALLBACK_RE = re.compile(r',\s*"([^"]+)"\)\s*$')


def extract_collection_id(raw_value: str) -> str:
    """Extract collection ID from a cell that may contain a Google Sheets IFERROR formula."""
    if raw_value is None:
        raise ValueError("Collection ID cell is empty")

    text = str(raw_value).strip()

    if "DUMMYFUNCTION" not in text:
        return text

    match = _FALLBACK_RE.search(text)
    if match:
        return match.group(1)

    raise ValueError(f"Cannot extract collection ID from formula: {text[-120:]}")


# ---------------------------------------------------------------------------
# Date expansion
# ---------------------------------------------------------------------------

def _parse_comma_list(value: Any) -> list[float] | None:
    """Parse a comma-separated string of numbers or a single float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return [float(value)]
    text = str(value).strip()
    if not text:
        return None
    return [float(x.strip()) for x in text.split(",")]


def _add_months(dt: datetime, months: int) -> datetime:
    """Add *months* calendar months to *dt*, clamping day to month length."""
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    # Clamp day
    import calendar
    max_day = calendar.monthrange(year, month)[1]
    day = min(dt.day, max_day)
    return dt.replace(year=year, month=month, day=day)


def _end_of_month(dt: datetime) -> datetime:
    """Return the last day of the month for *dt*."""
    import calendar
    last_day = calendar.monthrange(dt.year, dt.month)[1]
    return dt.replace(day=last_day)


def _end_of_year(dt: datetime) -> datetime:
    return dt.replace(month=12, day=31)


def expand_dates(
    start_date: datetime | None,
    end_date: datetime | None,
    date_step_raw: Any,
    date_offset_raw: Any,
    date_unit: str | None,
) -> list[tuple[datetime, datetime]]:
    """Return a list of (start, end) datetime pairs for the given parameters.

    Each pair represents one temporal slice to become a STAC item.
    """
    if start_date is None or end_date is None:
        return []

    unit = (date_unit or "").strip().lower()

    # ------------------------------------------------------------------
    # Case: static — single item spanning the full range
    # ------------------------------------------------------------------
    if unit == "static" or (date_step_raw is None and unit != "years" and unit != "months"):
        return [(start_date, end_date)]

    steps = _parse_comma_list(date_step_raw)
    offsets = _parse_comma_list(date_offset_raw)

    if steps is None:
        return [(start_date, end_date)]

    # ------------------------------------------------------------------
    # Case: uniform step (single value) with optional uniform offset
    # ------------------------------------------------------------------
    if len(steps) == 1:
        step = steps[0]
        offset = offsets[0] if offsets and len(offsets) == 1 else 0.0
        return _expand_uniform(start_date, end_date, step, offset, unit)

    # ------------------------------------------------------------------
    # Case: irregular steps (comma-separated list)
    # ------------------------------------------------------------------
    return _expand_irregular(start_date, end_date, steps, offsets, unit)


def _expand_uniform(
    start: datetime,
    end: datetime,
    step: float,
    offset: float,
    unit: str,
) -> list[tuple[datetime, datetime]]:
    """Expand with a uniform step size, optionally with a gap (offset) between intervals."""
    results: list[tuple[datetime, datetime]] = []
    cursor = start
    step_int = int(step)
    offset_int = int(offset)

    if unit == "years":
        while cursor <= end:
            interval_end = _end_of_year(
                cursor.replace(year=cursor.year + step_int - 1)
            )
            if interval_end > end:
                interval_end = end
            results.append((cursor, interval_end))
            # Advance by step + offset years
            cursor = cursor.replace(year=cursor.year + step_int + offset_int)
    elif unit == "months":
        while cursor <= end:
            interval_end = _add_months(cursor, step_int)
            # End of interval is day before next interval start
            interval_end = interval_end - timedelta(days=1)
            if interval_end > end:
                interval_end = end
            results.append((cursor, interval_end))
            cursor = _add_months(cursor, step_int + offset_int)
    else:
        # Fallback for unknown unit — treat as static
        results.append((start, end))

    return results


def _expand_irregular(
    start: datetime,
    end: datetime,
    steps: list[float],
    offsets: list[float] | None,
    unit: str,
) -> list[tuple[datetime, datetime]]:
    """Expand with per-interval step and offset lists.

    Each entry i uses steps[i] as the interval length and offsets[i] as a gap
    *before* the interval (year offset applied to the start cursor).
    """
    results: list[tuple[datetime, datetime]] = []
    cursor = start

    for i, step_val in enumerate(steps):
        if cursor > end:
            break

        step_int = int(step_val)
        year_offset = int(offsets[i]) if offsets and i < len(offsets) else 0

        if unit == "years":
            # Apply year offset to cursor (shift start of this interval)
            interval_start = cursor.replace(year=cursor.year + year_offset) if year_offset != 0 else cursor
            interval_end = _end_of_year(
                interval_start.replace(year=interval_start.year + step_int - 1)
            )
            if interval_end > end:
                interval_end = end
            results.append((interval_start, interval_end))
            # Next cursor: advance from interval_end's year + 1
            cursor = interval_end.replace(
                year=interval_end.year + 1, month=1, day=1
            )
        elif unit == "months":
            interval_start = cursor
            interval_end = _add_months(cursor, step_int) - timedelta(days=1)
            if interval_end > end:
                interval_end = end
            results.append((interval_start, interval_end))
            cursor = _add_months(cursor, step_int)
        else:
            results.append((start, end))
            break

    return results


# ---------------------------------------------------------------------------
# URL expansion
# ---------------------------------------------------------------------------

def _format_dt(start: datetime, end: datetime) -> str:
    """Format a date interval as YYYYMMdd_YYYYMMdd."""
    return f"{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}"


def expand_urls(
    main_url: str | None,
    extra_urls: dict[str, str],
    date_intervals: list[tuple[datetime, datetime]],
    depth_list: list[str],
) -> list[ItemDefinition]:
    """Generate one ItemDefinition per (date, depth) combination."""
    if main_url is None:
        return []

    has_dt = "{dt}" in main_url
    has_dp = "{dp}" in main_url

    depths = depth_list if (has_dp and depth_list) else [None]  # type: ignore[list-item]
    intervals = date_intervals if date_intervals else [(None, None)]  # type: ignore[list-item]

    items: list[ItemDefinition] = []

    for start_dt, end_dt in intervals:
        dt_str = _format_dt(start_dt, end_dt) if start_dt else ""

        for depth in depths:
            # Build item ID
            id_parts: list[str] = []
            if dt_str:
                id_parts.append(dt_str)
            if depth:
                id_parts.append(depth)
            item_id = "_".join(id_parts) if id_parts else "default"

            # Expand main URL
            href = main_url
            if has_dt:
                href = href.replace("{dt}", dt_str)
            if has_dp and depth:
                href = href.replace("{dp}", depth)

            assets: dict[str, str] = {"data": href}

            # Expand extra URLs
            for key, url in extra_urls.items():
                expanded = url
                if "{dt}" in expanded and dt_str:
                    expanded = expanded.replace("{dt}", dt_str)
                if "{dp}" in expanded and depth:
                    expanded = expanded.replace("{dp}", depth)
                assets[key] = expanded

            items.append(ItemDefinition(
                item_id=item_id,
                start_datetime=start_dt,
                end_datetime=end_dt,
                assets=assets,
            ))

    return items


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

# Column name -> 0-based index mapping (from header row)
_COL_MAP: dict[str, int] = {}


def _cell(ws: Any, row: int, col_name: str) -> Any:
    """Read cell by column name using the discovered header mapping."""
    idx = _COL_MAP[col_name]
    return ws.cell(row=row, column=idx + 1).value


def _build_col_map(ws: Any) -> None:
    """Build column name -> 0-based index mapping from the header row."""
    _COL_MAP.clear()
    for i, cell in enumerate(ws[1]):
        if cell.value:
            _COL_MAP[str(cell.value).strip()] = i


def parse_providers(wb: openpyxl.Workbook) -> dict[str, Provider]:
    """Parse the 'providers' sheet."""
    ws = wb["providers"]
    _build_col_map(ws)
    providers: dict[str, Provider] = {}

    for row_idx in range(2, ws.max_row + 1):
        name = _cell(ws, row_idx, "name")
        if not name:
            continue
        roles_raw = _cell(ws, row_idx, "roles_list") or ""
        providers[name] = Provider(
            name=name,
            description=_cell(ws, row_idx, "description") or "",
            roles=[r.strip() for r in str(roles_raw).split(",") if r.strip()],
            url=_cell(ws, row_idx, "url") or "",
        )

    return providers


def parse_catalog(wb: openpyxl.Workbook) -> CatalogInfo:
    """Parse the 'catalogs' sheet (single row expected)."""
    ws = wb["catalogs"]
    _build_col_map(ws)
    return CatalogInfo(
        id=_cell(ws, 2, "id") or "stac",
        title=_cell(ws, 2, "title") or "",
        description=_cell(ws, 2, "description") or "",
    )


def parse_collections(wb: openpyxl.Workbook) -> list[CollectionRecord]:
    """Parse the 'collections' sheet and expand all item definitions."""
    ws = wb["collections"]
    _build_col_map(ws)
    records: list[CollectionRecord] = []
    seen_ids: dict[str, int] = {}

    for row_idx in range(2, ws.max_row + 1):
        raw_id = _cell(ws, row_idx, "id")
        if raw_id is None:
            break

        collection_id = extract_collection_id(raw_id)

        # Handle duplicate IDs by appending a numeric suffix
        if collection_id in seen_ids:
            seen_ids[collection_id] += 1
            collection_id = f"{collection_id}_{seen_ids[collection_id]}"
        else:
            seen_ids[collection_id] = 0

        keywords_raw = _cell(ws, row_idx, "keywords_list") or ""
        keywords = [k.strip() for k in str(keywords_raw).split(",") if k.strip()]

        providers_raw = _cell(ws, row_idx, "providers_list") or ""
        provider_ids = [p.strip() for p in str(providers_raw).split(",") if p.strip()]

        depth_raw = _cell(ws, row_idx, "depth_list") or ""
        depth_list = [d.strip() for d in str(depth_raw).split(",") if d.strip()]

        # Extra URLs (url_1 .. url_13)
        extra_urls: dict[str, str] = {}
        for i in range(1, 14):
            col_name = f"url_{i}"
            if col_name in _COL_MAP:
                url_val = _cell(ws, row_idx, col_name)
                if url_val:
                    extra_urls[f"data_{i}"] = str(url_val)

        # SLD / QML style URLs
        sld_urls = []
        for col_name in ("sld_url_1", "sld_url_2"):
            if col_name in _COL_MAP:
                v = _cell(ws, row_idx, col_name)
                if v:
                    sld_urls.append(str(v))

        qml_urls = []
        for col_name in ("qml_url_1", "qml_url_2"):
            if col_name in _COL_MAP:
                v = _cell(ws, row_idx, col_name)
                if v:
                    qml_urls.append(str(v))

        start_date = _cell(ws, row_idx, "start_date")
        end_date = _cell(ws, row_idx, "end_date")
        date_step_raw = _cell(ws, row_idx, "date_step")
        date_offset_raw = _cell(ws, row_idx, "date_offset")
        date_unit = _cell(ws, row_idx, "date_unit")
        main_url = _cell(ws, row_idx, "main_url")

        # Expand dates
        date_intervals = expand_dates(
            start_date, end_date, date_step_raw, date_offset_raw, date_unit,
        )

        # Expand URLs into item definitions
        items = expand_urls(
            str(main_url) if main_url else None,
            extra_urls,
            date_intervals,
            depth_list,
        )

        record = CollectionRecord(
            id=collection_id,
            row_index=row_idx,
            dt_position=_cell(ws, row_idx, "dt_position"),
            catalog=_cell(ws, row_idx, "catalog") or "stac",
            collection_class=_cell(ws, row_idx, "class") or "",
            title=_cell(ws, row_idx, "title") or "",
            description=_cell(ws, row_idx, "description") or "",
            keywords=keywords,
            stac_url=_cell(ws, row_idx, "stac_url"),
            version=_cell(ws, row_idx, "version"),
            doi=_cell(ws, row_idx, "doi"),
            citation=_cell(ws, row_idx, "citation"),
            license=_cell(ws, row_idx, "license") or "proprietary",
            layer_unit=_cell(ws, row_idx, "layer_unit"),
            scale=_cell(ws, row_idx, "scale"),
            offset=_cell(ws, row_idx, "offset"),
            spatial_aggregation=_cell(ws, row_idx, "spatial_aggregation"),
            contact_name=_cell(ws, row_idx, "contact_name"),
            contact_email=_cell(ws, row_idx, "contact_email"),
            provider_ids=provider_ids,
            gsd=_cell(ws, row_idx, "gsd"),
            start_date=start_date,
            end_date=end_date,
            date_step=str(date_step_raw) if date_step_raw is not None else None,
            date_offset=str(date_offset_raw) if date_offset_raw is not None else None,
            date_unit=date_unit,
            date_style=_cell(ws, row_idx, "date_style"),
            depth_list=depth_list,
            main_url=str(main_url) if main_url else None,
            extra_urls=extra_urls,
            sld_urls=sld_urls,
            qml_urls=qml_urls,
            items=items,
        )
        records.append(record)

    return records


def parse_excel(excel_path: str | Path) -> ParseResult:
    """Parse the full Excel file and return all structured data."""
    wb = openpyxl.load_workbook(str(excel_path))
    catalog = parse_catalog(wb)
    providers = parse_providers(wb)
    collections = parse_collections(wb)

    total_items = sum(len(c.items) for c in collections)
    print(f"Parsed {len(collections)} collections, {total_items} items, "
          f"{len(providers)} providers")

    return ParseResult(
        catalog=catalog,
        providers=providers,
        collections=collections,
    )
